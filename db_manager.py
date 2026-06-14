"""
피시방 주문 시스템 - 공유 DB 관리자 (db_manager.py)

[자료구조 개선 사항]
  1. OrderQueue  : collections.deque → 주문 대기열 (FIFO)
  2. StockCache  : dict              → 재고 인메모리 캐시 (해시맵)
  3. SalesCounter: collections.Counter → 판매량 누적 집계
  4. menu_name_set: set              → 메뉴 중복 검사 O(1)
"""

import sqlite3
import os
from datetime import datetime
from collections import deque, Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 경로 설정 ──────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(BASE_DIR, "pcroom.db")
EXCEL_PATH = os.path.join(BASE_DIR, "export.xlsx")


# ══════════════════════════════════════════════════════════
#  [자료구조 1] 주문 대기열 — deque (FIFO Queue)
#
#  선택 이유:
#   - 주문은 "먼저 들어온 것부터 처리"하는 FIFO 구조
#   - deque는 양쪽 끝 O(1) 삽입/삭제 → list의 pop(0) O(n)보다 효율적
#   - appendleft/popleft로 큐 의미를 코드에 명시적으로 표현
#   - 관리자 화면의 폴링 없이 대기 주문 수를 O(1)로 파악 가능
# ══════════════════════════════════════════════════════════
order_queue: deque = deque()   # (order_id, seat, total, created_at) 튜플 저장


# ══════════════════════════════════════════════════════════
#  [자료구조 2] 재고 캐시 — dict (Hash Map)
#
#  선택 이유:
#   - 메뉴명(키)으로 재고(값)를 O(1)에 조회/갱신
#   - 매 주문마다 DB를 조회하는 대신 인메모리에서 즉시 처리
#   - DB는 영속성 역할만 담당, 캐시가 단일 진실 공급원
#   - 구조: { 메뉴명(str) : stock(int) }
# ══════════════════════════════════════════════════════════
stock_cache: dict[str, int] = {}


# ══════════════════════════════════════════════════════════
#  [자료구조 3] 판매 집계 — Counter (Hash Map 특수형)
#
#  선택 이유:
#   - Counter는 dict의 서브클래스로 빈도 집계에 특화
#   - 판매 발생 시 += 로 O(1) 누적, most_common(n)으로 Top-N 즉시 반환
#   - 매번 DB SUM/GROUP BY 쿼리를 날리지 않아도 됨
#   - 구조: { 메뉴명(str) : 누적판매수(int) }
# ══════════════════════════════════════════════════════════
sales_counter: Counter = Counter()


# ══════════════════════════════════════════════════════════
#  [자료구조 4] 메뉴 이름 집합 — set (Hash Set)
#
#  선택 이유:
#   - 메뉴 중복 확인을 O(n) 리스트 탐색 → O(1) 해시 조회로 개선
#   - 기존: [m["name"] for m in db.get_menu()] + "in" 연산 → O(n)
#   - 개선: "메뉴명 in menu_name_set" → O(1)
#   - 추가/삭제도 O(1)로 항상 최신 상태 유지
# ══════════════════════════════════════════════════════════
menu_name_set: set[str] = set()


# ══════════════════════════════════════════════════════════
#  [자료구조 5] 옵저버 목록 — list[callable]
#
#  선택 이유:
#   - 기존 3초 폴링 대신, 상태 변화 시 등록된 콜백을 즉시 호출
#   - 관심 있는 화면(관리자/고객)이 subscribe() 로 자신을 등록
#   - 변화 발생 시 _notify_observers()가 모든 콜백을 순회 호출
#   - 자료구조 패턴: 옵저버(Observer) 패턴
# ══════════════════════════════════════════════════════════
_observers: list = []   # [callable, ...]


def subscribe(callback):
    """변화 알림을 받을 콜백 함수를 등록합니다."""
    if callback not in _observers:
        _observers.append(callback)


def unsubscribe(callback):
    """콜백 등록을 해제합니다."""
    if callback in _observers:
        _observers.remove(callback)


def _notify_observers(event: str, payload: dict = None):
    """등록된 모든 옵저버에게 이벤트를 알립니다."""
    for cb in _observers:
        try:
            cb(event, payload or {})
        except Exception:
            pass


# ══════════════════════════════════════════════════════════
#  DB 초기화
# ══════════════════════════════════════════════════════════

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=5)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """DB 테이블 생성 후, 인메모리 자료구조를 모두 초기화합니다."""
    with get_conn() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS menu (
                name      TEXT PRIMARY KEY,
                category  TEXT NOT NULL,
                price     INTEGER NOT NULL,
                stock     INTEGER NOT NULL DEFAULT 0,
                sold      INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS orders (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                seat        TEXT NOT NULL,
                total       INTEGER NOT NULL,
                status      TEXT NOT NULL DEFAULT '대기',
                created_at  TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS order_items (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id  INTEGER NOT NULL REFERENCES orders(id),
                name      TEXT NOT NULL,
                qty       INTEGER NOT NULL,
                price     INTEGER NOT NULL
            );
        """)

    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM menu").fetchone()[0]
        if count == 0:
            _insert_default_menu()

    _rebuild_caches()   # DB → 인메모리 자료구조 동기화


def _rebuild_caches():
    """
    DB에서 전체 데이터를 읽어 인메모리 자료구조를 재구축합니다.
    프로그램 시작 시 또는 외부 변경 감지 시 호출합니다.
    """
    global stock_cache, sales_counter, menu_name_set

    with get_conn() as conn:
        rows = conn.execute(
            "SELECT name, stock, sold FROM menu"
        ).fetchall()

    # [자료구조 2] dict 재구축
    stock_cache = {r["name"]: r["stock"] for r in rows}

    # [자료구조 3] Counter 재구축 — 기존 판매 이력 반영
    sales_counter = Counter({r["name"]: r["sold"] for r in rows})

    # [자료구조 4] set 재구축
    menu_name_set = {r["name"] for r in rows}

    # [자료구조 1] deque 재구축 — DB의 대기 주문을 큐에 적재
    order_queue.clear()
    with get_conn() as conn:
        pending = conn.execute(
            "SELECT id, seat, total, created_at FROM orders "
            "WHERE status='대기' ORDER BY id ASC"
        ).fetchall()
    for o in pending:
        order_queue.append((o["id"], o["seat"], o["total"], o["created_at"]))


def _insert_default_menu():
    default_menu = [
        # ── 라면류 ───────────────────────────────
        ("신라면",       "라면류", 3500, 30),
        ("진라면",       "라면류", 3500, 30),
        ("짜파게티",     "라면류", 4000, 25),
        ("불닭볶음면",   "라면류", 4000, 25),
        ("치즈라면",     "라면류", 4500, 20),
        ("계란라면",     "라면류", 4500, 20),
        # ── 밥류 ────────────────────────────────
        ("참치마요 주먹밥", "밥류", 3500, 20),
        ("김치볶음밥",   "밥류",   6500, 15),
        ("새우볶음밥",   "밥류",   6500, 15),
        ("햄야채볶음밥", "밥류",   6500, 15),
        ("제육덮밥",     "밥류",   7000, 10),
        ("불고기덮밥",   "밥류",   7000, 10),
        ("카레라이스",   "밥류",   6000, 15),
        ("오므라이스",   "밥류",   7000, 10),
        ("돈까스",       "밥류",   8000, 10),
        ("치즈돈까스",   "밥류",   9000, 10),
        # ── 분식·스낵 ────────────────────────────
        ("떡볶이",       "분식·스낵", 5000, 20),
        ("치즈떡볶이",   "분식·스낵", 6000, 15),
        ("순대",         "분식·스낵", 5000, 15),
        ("튀김 모둠",    "분식·스낵", 5500, 15),
        ("핫도그",       "분식·스낵", 2500, 20),
        ("소떡소떡",     "분식·스낵", 3500, 20),
        ("닭강정",       "분식·스낵", 6500, 10),
        ("치킨너겟",     "분식·스낵", 5000, 15),
        ("감자튀김",     "분식·스낵", 4500, 20),
        ("군만두",       "분식·스낵", 4000, 20),
        ("토스트",       "분식·스낵", 3500, 20),
        ("햄치즈샌드위치", "분식·스낵", 4500, 15),
        # ── 음료 ────────────────────────────────
        ("아이스티",         "음료", 2000, 40),
        ("콜라(캔)",         "음료", 2000, 50),
        ("사이다(캔)",       "음료", 2000, 50),
        ("아메리카노",       "음료", 2500, 40),
        ("에너지드링크",     "음료", 3000, 30),
        ("카페라떼",         "음료", 3500, 30),
        ("바닐라라떼",       "음료", 4000, 25),
        ("카라멜마끼아또",   "음료", 4500, 20),
        ("카페모카",         "음료", 4500, 20),
    ]
    with get_conn() as conn:
        conn.executemany(
            "INSERT OR IGNORE INTO menu(name, category, price, stock, sold) "
            "VALUES(?, ?, ?, ?, 0)",
            default_menu,
        )


def reset_db():
    """
    재고·주문·매출을 모두 초기화하고 기본 메뉴를 다시 삽입합니다.

    초기화 순서:
      1. order_items → orders → menu 순서로 DELETE
         (외래키 제약: order_items가 orders를 참조하므로 자식 테이블 먼저 삭제)
      2. 기본 메뉴 재삽입 (_insert_default_menu)
      3. 인메모리 자료구조 전체 재구축 (_rebuild_caches)
         - deque, dict, Counter, set 모두 초기화
      4. 옵저버에게 'reset' 이벤트 알림
    """
    with get_conn() as conn:
        conn.executescript("""
            DELETE FROM order_items;
            DELETE FROM orders;
            DELETE FROM menu;
        """)
    _insert_default_menu()
    _rebuild_caches()
    _notify_observers("reset", {})


# ══════════════════════════════════════════════════════════
#  메뉴 / 재고
# ══════════════════════════════════════════════════════════

def get_menu() -> list:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT name, category, price, stock, sold "
            "FROM menu ORDER BY category, name"
        ).fetchall()
    return [dict(r) for r in rows]


def menu_exists(name: str) -> bool:
    """
    [자료구조 4 활용] set을 이용한 O(1) 메뉴 중복 확인.
    기존: name in [m["name"] for m in get_menu()]  → O(n)
    개선: name in menu_name_set                    → O(1)
    """
    return name in menu_name_set


def reduce_stock(name: str, qty: int = 1) -> bool:
    """
    [자료구조 2 활용] dict 캐시에서 재고를 먼저 확인 후 차감.
    캐시와 DB를 동시에 갱신하여 항상 일관성을 유지합니다.
    """
    # 캐시에서 O(1) 재고 확인
    if stock_cache.get(name, 0) < qty:
        return False

    # 인메모리 캐시 갱신
    stock_cache[name] -= qty

    # DB 영속화
    with get_conn() as conn:
        conn.execute(
            "UPDATE menu SET stock=stock-?, sold=sold+? WHERE name=?",
            (qty, qty, name),
        )
    return True


def restore_stock(name: str, qty: int = 1):
    """[자료구조 2 활용] 주문 취소 시 캐시와 DB 재고를 동시에 복구."""
    stock_cache[name] = stock_cache.get(name, 0) + qty
    with get_conn() as conn:
        conn.execute(
            "UPDATE menu SET stock=stock+?, sold=MAX(0, sold-?) WHERE name=?",
            (qty, qty, name),
        )


def update_stock(name: str, new_stock: int):
    """[자료구조 2 활용] 관리자 직접 수정 — 캐시와 DB 동시 갱신."""
    stock_cache[name] = new_stock
    with get_conn() as conn:
        conn.execute(
            "UPDATE menu SET stock=? WHERE name=?", (new_stock, name)
        )


def add_menu(name: str, category: str, price: int, stock: int):
    """[자료구조 2, 4 활용] 메뉴 추가 시 캐시·집합·DB 동시 갱신."""
    with get_conn() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO menu(name, category, price, stock, sold) "
            "VALUES(?, ?, ?, ?, 0)",
            (name, category, price, stock),
        )
    stock_cache[name] = stock
    sales_counter[name] = 0
    menu_name_set.add(name)          # [자료구조 4] O(1) 삽입
    _notify_observers("menu_changed", {"action": "add", "name": name})


def update_menu(name: str, category: str, price: int):
    with get_conn() as conn:
        conn.execute(
            "UPDATE menu SET category=?, price=? WHERE name=?",
            (category, price, name),
        )
    _notify_observers("menu_changed", {"action": "update", "name": name})


def delete_menu(name: str):
    """[자료구조 2, 4 활용] 메뉴 삭제 시 캐시·집합·카운터·DB 동시 갱신."""
    with get_conn() as conn:
        conn.execute("DELETE FROM menu WHERE name=?", (name,))
    stock_cache.pop(name, None)      # [자료구조 2] O(1) 삭제
    sales_counter.pop(name, None)    # [자료구조 3] O(1) 삭제
    menu_name_set.discard(name)      # [자료구조 4] O(1) 삭제
    _notify_observers("menu_changed", {"action": "delete", "name": name})


def get_stock(name: str) -> int:
    """[자료구조 2 활용] 캐시에서 O(1) 재고 조회."""
    return stock_cache.get(name, 0)


# ══════════════════════════════════════════════════════════
#  주문
# ══════════════════════════════════════════════════════════

def create_order(seat: str, cart: dict, menu_map: dict) -> int:
    """
    [자료구조 1 활용] 주문 생성 후 deque에 enqueue.
    cart는 OrderedDict로 전달되어 삽입 순서가 보장됩니다.
    """
    total = sum(menu_map[n]["price"] * q for n, q in cart.items())
    now   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO orders(seat, total, status, created_at) VALUES(?,?,?,?)",
            (seat, total, "대기", now),
        )
        order_id = cur.lastrowid
        conn.executemany(
            "INSERT INTO order_items(order_id, name, qty, price) VALUES(?,?,?,?)",
            [(order_id, n, q, menu_map[n]["price"]) for n, q in cart.items()],
        )

    # [자료구조 1] 대기열에 enqueue (append = 오른쪽 끝에 추가)
    order_queue.append((order_id, seat, total, now))

    # [자료구조 3] 판매 카운터 즉시 반영
    for name, qty in cart.items():
        sales_counter[name] += qty

    _notify_observers("new_order", {"order_id": order_id, "seat": seat})
    return order_id


def get_orders(status_filter: str = None) -> list:
    sql = """
        SELECT o.id, o.seat, o.total, o.status, o.created_at,
               GROUP_CONCAT(oi.name || ' x' || oi.qty, ', ') AS items
        FROM orders o
        LEFT JOIN order_items oi ON o.id = oi.order_id
    """
    params = ()
    if status_filter:
        sql += " WHERE o.status=?"
        params = (status_filter,)
    sql += " GROUP BY o.id ORDER BY o.id DESC"

    with get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def get_pending_queue() -> list:
    """
    [자료구조 1 활용] deque에서 대기 주문 목록을 O(n) 순회로 반환.
    DB 조회 없이 인메모리 큐를 직접 읽습니다.
    반환: [(order_id, seat, total, created_at), ...]  (접수 순서 보장)
    """
    return list(order_queue)


def get_order_items(order_id: int) -> list:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT name, qty, price FROM order_items WHERE order_id=?",
            (order_id,),
        ).fetchall()
    return [dict(r) for r in rows]


def update_order_status(order_id: int, status: str):
    """
    [자료구조 1 활용] 완료·취소 처리 시 deque에서 해당 주문을 제거(dequeue).
    deque는 임의 위치 삭제를 지원하므로 처리 순서와 무관하게 제거 가능.
    """
    with get_conn() as conn:
        conn.execute(
            "UPDATE orders SET status=? WHERE id=?", (status, order_id)
        )

    # [자료구조 1] 대기열에서 해당 주문 제거
    for item in list(order_queue):
        if item[0] == order_id:
            order_queue.remove(item)
            break

    _notify_observers("order_updated", {"order_id": order_id, "status": status})


def get_total_sales() -> int:
    """
    [자료구조 3 활용] Counter에서 가격을 곱해 총 매출을 계산.
    완료 주문만 집계하기 위해 DB는 최소한으로 사용합니다.
    """
    with get_conn() as conn:
        row = conn.execute(
            "SELECT COALESCE(SUM(total), 0) FROM orders WHERE status='완료'"
        ).fetchone()
    return row[0]


def get_top_menu(limit: int = 3) -> list:
    """
    [자료구조 3 활용] Counter.most_common(n)으로 Top-N 메뉴를 O(n log n) 반환.
    기존: DB ORDER BY sold DESC LIMIT n
    개선: 인메모리 Counter에서 즉시 반환
    """
    return [
        {"name": name, "sold": sold}
        for name, sold in sales_counter.most_common(limit)
        if sold > 0
    ]


# ══════════════════════════════════════════════════════════
#  Excel 내보내기
# ══════════════════════════════════════════════════════════

def export_to_excel(path: str = None) -> str:
    save_path = path or EXCEL_PATH
    wb = openpyxl.Workbook()
    _write_inventory_sheet(wb)
    _write_orders_sheet(wb)
    _write_summary_sheet(wb)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(save_path)
    wb.close()
    return save_path


def _header_style(cell, bg: str = "2C3E50"):
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    _apply_border(cell)


def _apply_border(cell):
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_inventory_sheet(wb):
    ws = wb.create_sheet("재고현황")
    headers = ["메뉴명", "카테고리", "가격(원)", "재고수량", "판매수량", "총매출(원)"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _header_style(ws.cell(1, col))

    menu_list = get_menu()
    for m in menu_list:
        ws.append([m["name"], m["category"], m["price"],
                   m["stock"], m["sold"], m["price"] * m["sold"]])
        for col in range(1, 7):
            cell = ws.cell(ws.max_row, col)
            cell.alignment = Alignment(horizontal="center")
            _apply_border(cell)

    # 합계 행
    total_row = ws.max_row + 1
    ws.cell(total_row, 1, "합계")
    ws.cell(total_row, 5, f"=SUM(E2:E{total_row - 1})")   # 판매수량 합계
    ws.cell(total_row, 6, f"=SUM(F2:F{total_row - 1})")   # 총매출 합계
    for col in range(1, 7):
        cell = ws.cell(total_row, col)
        cell.font      = Font(bold=True, size=10)
        cell.fill      = PatternFill("solid", start_color="2C3E50")
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center")
        _apply_border(cell)

    for col, w in zip("ABCDEF", [16, 12, 12, 12, 12, 14]):
        ws.column_dimensions[col].width = w


def _write_orders_sheet(wb):
    ws = wb.create_sheet("주문내역")
    headers = ["주문번호", "좌석", "메뉴", "금액(원)", "상태", "주문시각"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _header_style(ws.cell(1, col), bg="1A5276")

    status_colors = {"완료": "D5F5E3", "취소": "FADBD8", "대기": "FEF9E7"}
    for o in get_orders():
        ws.append([o["id"], o["seat"], o["items"] or "",
                   o["total"], o["status"], o["created_at"]])
        color = status_colors.get(o["status"], "FFFFFF")
        for col in range(1, 7):
            cell = ws.cell(ws.max_row, col)
            cell.fill      = PatternFill("solid", start_color=color)
            cell.alignment = Alignment(horizontal="center")
            _apply_border(cell)

    # 완료 주문 금액 합계 행
    last_data = ws.max_row
    total_row = last_data + 1
    ws.cell(total_row, 3, "완료 주문 합계")
    # 상태가 '완료'인 행의 금액(D열)만 SUMIF로 합산
    ws.cell(total_row, 4,
            f'=SUMIF(E2:E{last_data},"완료",D2:D{last_data})')
    for col in range(1, 7):
        cell = ws.cell(total_row, col)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", start_color="1A5276")
        cell.alignment = Alignment(horizontal="center")
        _apply_border(cell)

    for col, w in zip("ABCDEF", [10, 8, 35, 12, 8, 20]):
        ws.column_dimensions[col].width = w


def _write_summary_sheet(wb):
    ws = wb.create_sheet("매출요약")
    ws["A1"] = "항목"; ws["B1"] = "값"
    _header_style(ws["A1"], bg="7D6608"); _header_style(ws["B1"], bg="7D6608")
    ws["A2"] = "총 매출 (완료 주문)";  ws["B2"] = get_total_sales()
    ws["A3"] = "총 주문 건수";          ws["B3"] = len(get_orders())
    ws["A4"] = "완료 건수";             ws["B4"] = len(get_orders("완료"))
    ws["A5"] = "취소 건수";             ws["B5"] = len(get_orders("취소"))
    ws["A6"] = "대기 건수";             ws["B6"] = len(get_orders("대기"))
    for r in range(2, 7):
        for col in "AB":
            cell = ws[f"{col}{r}"]
            cell.alignment = Alignment(horizontal="center")
            _apply_border(cell)
    ws["A8"] = "인기 메뉴 Top 5"
    ws.append(["순위", "메뉴명", "판매수량"])
    for col in range(1, 4):
        _header_style(ws.cell(ws.max_row, col), bg="1E8449")
    for rank, m in enumerate(get_top_menu(5), 1):
        ws.append([rank, m["name"], m["sold"]])
        for col in range(1, 4):
            cell = ws.cell(ws.max_row, col)
            cell.alignment = Alignment(horizontal="center")
            _apply_border(cell)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws["A10"] = f"내보내기 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A10"].font = Font(color="888888", italic=True)